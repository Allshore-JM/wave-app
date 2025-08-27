from flask import Flask, render_template, request, send_file, jsonify
import pandas as pd  # still used for fallback Excel generation if needed
import requests
import json
import os
from io import BytesIO
from datetime import datetime, timedelta
import pytz
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side
from timezonefinder import TimezoneFinder
import re

app = Flask(__name__)

# Instantiate a timezone finder object once to avoid repeated initialization.
tz_finder = TimezoneFinder()

# Caches for station metadata and available bulletin stations. These will be populated
# on demand to avoid repeated network requests. See load_station_metadata() and
# get_bullet_station_ids().
STATION_META = None  # Maps station_id -> { 'name': str, 'lat': float, 'lon': float }

# -----------------------------------------------------------------------------
# Coordinate database
#
# To avoid making network requests for every buoy when rendering the map, the
# application can load precomputed latitude/longitude values for each buoy
# station from a static JSON file.  The file ``station_coords.json`` is
# generated offline (for example from the provided Excel spreadsheet) and
# placed alongside this module.  When present, it contains an object mapping
# station IDs to ``{"lat": float, "lon": float}``.  This allows the map
# to display all stations immediately without having to fetch .bull files.
STATION_COORDS = None  # Maps station_id -> { 'lat': float, 'lon': float }

def load_station_coords() -> dict:
    """
    Load the static station coordinates mapping from ``station_coords.json``.
    """
    global STATION_COORDS
    if STATION_COORDS is not None:
        return STATION_COORDS
    base_dir = os.path.dirname(os.path.abspath(__file__))
    coord_path = os.path.join(base_dir, 'station_coords.json')
    coords = {}
    try:
        with open(coord_path, 'r') as f:
            data = json.load(f)
            # Ensure values are floats
            for sid, info in data.items():
                try:
                    lat = float(info.get('lat'))
                    lon = float(info.get('lon'))
                    coords[str(sid).strip()] = {'lat': lat, 'lon': lon}
                except Exception:
                    continue
    except Exception:
        coords = {}
    STATION_COORDS = coords
    return STATION_COORDS

# -----------------------------------------------------------------------------
# Default station metadata fallback
#
# When the application is unable to retrieve the station list or metadata from
# the NOAA servers (for example due to network restrictions), we fall back to
# a small curated set of buoy stations.  Each entry provides an approximate
# latitude/longitude and a human‑readable name.  This ensures that the map
# displays multiple markers and the dropdown contains more than a single
# placeholder buoy even when remote resources are unavailable.  You can add
# additional stations here as desired.  Latitude values are positive for
# northern hemisphere and negative for southern; longitude values are negative
# for western hemisphere and positive for eastern.
DEFAULT_STATIONS = {
    # Hawaiian region buoys
    "51201": {"name": "Buoy 51201", "lat": 21.67, "lon": -158.12},
    "51202": {"name": "Buoy 51202", "lat": 21.45, "lon": -157.90},
    "51203": {"name": "Buoy 51203", "lat": 21.55, "lon": -157.95},
    "51211": {"name": "Buoy 51211", "lat": 21.32, "lon": -157.53},
    "51212": {"name": "Buoy 51212", "lat": 21.27, "lon": -157.47},
    "51213": {"name": "Buoy 51213", "lat": 21.17, "lon": -157.17},
    # Other North Pacific buoys
    "51001": {"name": "Buoy 51001", "lat": 16.87, "lon": -156.47},
    "51002": {"name": "Buoy 51002", "lat": 12.38, "lon": -157.49},
    "51003": {"name": "Buoy 51003", "lat": 23.69, "lon": -162.25},
    "51004": {"name": "Buoy 51004", "lat": 25.84, "lon": -162.09},
}
BULLET_STATIONS = None  # Set of station_ids that currently have a .bull file available

# ===== NOAA Station List URL =====
# This URL points to the bulls.readme file, which lists the available buoy stations
STATION_LIST_URL = "https://nomads.ncep.noaa.gov/pub/data/nccf/com/gfs/prod/wave/station/bulls.readme"

# ===== Base NOAA URL Pattern =====
# The pattern for locating GFS wave data. The date (YYYYMMDD) and run hour (HH) will be inserted.
NOAA_BASE = "https://nomads.ncep.noaa.gov/pub/data/nccf/com/gfs/prod"

# ===== Timezones =====
# The app displays times in UTC and Hawaii Standard Time (HST). HST is used instead of the local timezone because many users
# of buoy data in Hawaii prefer local time display. The user’s timezone is configured via pytz.
HST = pytz.timezone("Pacific/Honolulu")
UTC = pytz.utc

# Cache for compiled station metadata used by the interactive map.  This list is
# computed once on demand and reused for subsequent requests.  Each element
# contains the station ID, name, latitude and longitude.
stations_data_cache = None


def get_station_list() -> list[tuple[str, str]]:
    """
    Build a list of available stations for the dropdown using the static
    ``station_list.json`` file.
    """
    stations: list[tuple[str, str]] = []
    try:
        base_dir = os.path.dirname(os.path.abspath(__file__))
        json_path = os.path.join(base_dir, 'station_list.json')
        with open(json_path, 'r') as f:
            station_ids = json.load(f)
        try:
            meta = load_station_metadata()
        except Exception:
            meta = {}
        for sid in station_ids:
            sid_str = str(sid).strip()
            if not sid_str:
                continue
            info = meta.get(sid_str)
            name = info['name'] if info and 'name' in info else sid_str
            stations.append((sid_str, name))
        if stations:
            return stations
    except Exception:
        pass
    return [(sid, info.get('name', sid)) for sid, info in DEFAULT_STATIONS.items()]

def get_stations_data():
    """
    Build and return a list of station metadata dictionaries: {id, name, lat, lon}.
    """
    global stations_data_cache
    if stations_data_cache is not None:
        return stations_data_cache
    try:
        ids_with_names = get_station_list()
        id_list = [sid for sid, _ in ids_with_names]
    except Exception:
        id_list = []
    try:
        meta = load_station_metadata()
    except Exception:
        meta = {}
    coords_map = load_station_coords()
    data_list = []
    for sid in id_list:
        name = sid
        info = meta.get(sid)
        if info and 'name' in info:
            name = info['name']
        lat = lon = None
        if sid in coords_map:
            lat = coords_map[sid]['lat']
            lon = coords_map[sid]['lon']
        if (lat is None or lon is None) and sid in DEFAULT_STATIONS:
            fallback_info = DEFAULT_STATIONS[sid]
            lat = fallback_info.get('lat')
            lon = fallback_info.get('lon')
        if lat is not None and lon is not None:
            data_list.append({
                'id': sid,
                'name': name,
                'lat': lat,
                'lon': lon,
            })
    stations_data_cache = data_list
    return stations_data_cache

@app.route('/stations.json')
def stations_json():
    """
    JSON endpoint returning the list of station metadata dictionaries used by
    the front‑end map.
    """
    return jsonify(get_stations_data())


def load_station_metadata():
    """
    Load buoy station metadata including latitude, longitude and station name from the
    NDBC station table.
    """
    global STATION_META
    if STATION_META is not None:
        return STATION_META
    station_url = "https://www.ndbc.noaa.gov/data/stations/station_table.txt"
    meta = {}
    try:
        res = requests.get(station_url, timeout=30)
        res.raise_for_status()
        for line in res.text.splitlines():
            if not line or line.startswith('#'):
                continue
            parts = line.split('|')
            if len(parts) < 7:
                continue
            station_id = parts[0].strip()
            if not station_id:
                continue
            name = parts[4].strip() if parts[4].strip() else station_id
            location_field = parts[6].strip()
            tokens = location_field.split()
            if len(tokens) >= 4:
                try:
                    lat_val = float(tokens[0])
                    lat_dir = tokens[1].upper()
                    lon_val = float(tokens[2])
                    lon_dir = tokens[3].upper()
                    lat = lat_val if lat_dir == 'N' else -lat_val
                    lon = lon_val if lon_dir == 'E' else -lon_val
                    meta[station_id] = {'name': name, 'lat': lat, 'lon': lon}
                except Exception:
                    continue
        STATION_META = meta
        return STATION_META
    except Exception:
        STATION_META = DEFAULT_STATIONS.copy()
        return STATION_META


def get_bullet_station_ids():
    """
    Retrieve the set of station IDs for which a GFS wave bulletin is currently available.
    """
    global BULLET_STATIONS
    if BULLET_STATIONS is not None:
        return BULLET_STATIONS
    date_str, run_str = get_latest_run()
    if not date_str:
        BULLET_STATIONS = set()
        return BULLET_STATIONS
    url = f"{NOAA_BASE}/gfs.{date_str}/{run_str}/wave/station/bulls.t{run_str}z/"
    ids = set()
    try:
        res = requests.get(url, timeout=30)
        res.raise_for_status()
        from bs4 import BeautifulSoup
        soup = BeautifulSoup(res.text, "html.parser")
        for link in soup.find_all('a'):
            href = link.get('href', '')
            if href.startswith('gfswave.') and href.endswith('.bull'):
                parts = href.split('.')
                if len(parts) >= 2:
                    sid = parts[1]
                    ids.add(sid)
        BULLET_STATIONS = ids
        return BULLET_STATIONS
    except Exception:
        BULLET_STATIONS = set()
        return BULLET_STATIONS


def get_latest_run():
    """
    Determine the most recent available GFS model run.
    """
    now = datetime.utcnow()
    run_hours = [18, 12, 6, 0]
    for delta_day in [0, 1]:
        check_date = now - timedelta(days=delta_day)
        yyyymmdd = check_date.strftime("%Y%m%d")
        for hour in run_hours:
            run_str = f"{hour:02d}"
            url = f"{NOAA_BASE}/gfs.{yyyymmdd}/{run_str}/wave/station/bulls.t{run_str}z/"
            test_file = f"{url}gfswave.51201.bull"
            try:
                resp = requests.head(test_file, timeout=10)
            except Exception:
                continue
            if resp.status_code == 200:
                return yyyymmdd, run_str
    return None, None


def parse_bull(station_id: str, target_tz_name: str | None = None):
    """
    Fetch and parse the .bull file for a given station.
    Returns: (cycle_str, location_str, model_run_str, rows, tz_name, error)
    """
    date_str, run_str = get_latest_run()
    if not date_str:
        return None, None, None, None, 'UTC', "No recent run found."

    bull_url = f"{NOAA_BASE}/gfs.{date_str}/{run_str}/wave/station/bulls.t{run_str}z/gfswave.{station_id}.bull"
    try:
        resp = requests.get(bull_url, timeout=10)
    except Exception as e:
        return None, None, None, None, 'UTC', f"Error retrieving .bull file: {e}"
    if resp.status_code != 200:
        return None, None, None, None, 'UTC', f"No .bull file found for {station_id}"

    lines = resp.text.splitlines()
    if not lines:
        return None, None, None, None, 'UTC', "Downloaded .bull file is empty."

    cycle_line = next((l for l in lines if l.lower().strip().startswith("cycle")), None)
    location_line = next((l for l in lines if l.lower().strip().startswith("location")), None)
    if not cycle_line and len(lines) > 0:
        cycle_line = lines[0]
    if not location_line and len(lines) > 1:
        location_line = lines[1]
    cycle_str = cycle_line.strip() if cycle_line else ""
    location_str = location_line.strip() if location_line else ""

    # Extract lat/lon from header if present
    lat = lon = None
    tz_name = 'UTC'
    if location_str:
        m = re.search(r"\(([-+]?\d+(?:\.\d+)?)\s*([NS])\s+([-+]?\d+(?:\.\d+)?)\s*([EW])\)", location_str)
        if m:
            try:
                lat_val = float(m.group(1))
                lat_dir = m.group(2)
                lon_val = float(m.group(3))
                lon_dir = m.group(4)
                lat = lat_val if lat_dir.upper() == 'N' else -lat_val
                lon = lon_val if lon_dir.upper() == 'E' else -lon_val
            except Exception:
                lat = lon = None
    if lat is not None and lon is not None:
        try:
            tz_name_candidate = tz_finder.timezone_at(lat=lat, lng=lon)
            if tz_name_candidate:
                tz_name = tz_name_candidate
        except Exception:
            tz_name = 'UTC'

    effective_tz_name = tz_name
    if target_tz_name:
        try:
            _ = pytz.timezone(target_tz_name)
            effective_tz_name = target_tz_name
        except Exception:
            pass

    # Detect file format
    uses_day_hour_format = any("day &" in line.lower() for line in lines[:10])

    rows = []

    if uses_day_hour_format:
        # ----- Newer format parser: 'day & hour' -----
        m = re.search(r"(\d{8})\s*(\d{2})", cycle_str)
        cycle_date_str = date_str
        cycle_hour_str = run_str
        if m:
            cycle_date_str = m.group(1)
            cycle_hour_str = m.group(2)
        try:
            cycle_dt_utc = datetime.strptime(f"{cycle_date_str} {cycle_hour_str}", "%Y%m%d %H")
        except Exception:
            cycle_dt_utc = datetime.strptime(f"{date_str} {run_str}", "%Y%m%d %H")

        try:
            model_run_local = cycle_dt_utc.replace(tzinfo=UTC).astimezone(pytz.timezone(effective_tz_name))
        except Exception:
            model_run_local = cycle_dt_utc
        try:
            model_run_str = "Model Run: " + model_run_local.strftime("%A, %B %-d, %Y %I:%M %p")
        except Exception:
            model_run_str = "Model Run: " + model_run_local.strftime("%A, %B %d, %Y %I:%M %p").lstrip('0')

        # Helper: find earliest datetime >= threshold with requested (day_of_month, hour)
        def _first_dt_on_or_after(threshold_dt: datetime, day_val: int, hour_val: int) -> datetime | None:
            base_year = threshold_dt.year
            base_month = threshold_dt.month
            for offset in range(0, 14):
                y = base_year + ((base_month - 1 + offset) // 12)
                mth = ((base_month - 1 + offset) % 12) + 1
                try:
                    cand = datetime(y, mth, day_val, hour_val)
                except ValueError:
                    continue
                if cand >= threshold_dt:
                    return cand
            return None

        M_TO_FT = 3.28084
        prev_forecast_dt_utc: datetime | None = None

        for line in lines:
            striped = line.strip()
            if not striped.startswith("|"):
                continue
            if "Hst" in striped or "---" in striped:
                continue

            parts = [p.strip() for p in line.split("|") if p.strip()]
            if not parts:
                continue

            day_hour = parts[0].split()
            if len(day_hour) < 2:
                continue
            try:
                day_val = int(day_hour[0])
                hour_val = int(day_hour[1])
            except ValueError:
                continue

            hst_tokens = parts[1].split()
            combined_hs_m = None
            if hst_tokens:
                try:
                    combined_hs_m = float(hst_tokens[0].replace('*', ''))
                except ValueError:
                    combined_hs_m = None

            swell_groups = []
            for swell_field in parts[2:]:
                if not swell_field:
                    swell_groups.append((None, None, None))
                    continue
                raw_tokens = swell_field.split()
                cleaned_tokens = []
                for tok in raw_tokens:
                    tok_clean = tok.replace('*', '')
                    if tok_clean == '':
                        continue
                    cleaned_tokens.append(tok_clean)
                if len(cleaned_tokens) < 3:
                    swell_groups.append((None, None, None))
                else:
                    try:
                        hs_val = float(cleaned_tokens[0])
                        tp_val = float(cleaned_tokens[1])
                        dir_raw = int(float(cleaned_tokens[2]))
                        dir_val = (dir_raw + 180) % 360
                        swell_groups.append((hs_val, tp_val, dir_val))
                    except ValueError:
                        swell_groups.append((None, None, None))

            while len(swell_groups) < 6:
                swell_groups.append((None, None, None))
            if len(swell_groups) > 6:
                swell_groups = swell_groups[:6]

            # Month-end continuity: never go backward in time
            threshold_dt = prev_forecast_dt_utc if (prev_forecast_dt_utc and prev_forecast_dt_utc > cycle_dt_utc) else cycle_dt_utc
            forecast_dt_utc = _first_dt_on_or_after(threshold_dt, day_val, hour_val)
            if forecast_dt_utc is None:
                continue
            prev_forecast_dt_utc = forecast_dt_utc

            try:
                local_tz = pytz.timezone(effective_tz_name)
            except Exception:
                local_tz = UTC
            local_dt = forecast_dt_utc.replace(tzinfo=UTC).astimezone(local_tz)
            try:
                date_str_local = local_dt.strftime("%A, %B %-d, %Y")
            except Exception:
                date_str_local = local_dt.strftime("%A, %B %d, %Y").lstrip('0')
            time_str_local = local_dt.strftime("%I:%M %p").lstrip('0')

            combined_hs_ft = None
            if combined_hs_m is not None:
                combined_hs_ft = combined_hs_m * M_TO_FT

            row = [date_str_local, time_str_local]
            for hs_m, tp_val, dir_val in swell_groups:
                if hs_m is None:
                    row.extend([None, None, None])
                else:
                    row.extend([hs_m * M_TO_FT, tp_val, dir_val])
            row.append(combined_hs_ft)
            rows.append(row)

    else:
        # ----- Older format parser: header contains "Hr" followed by swell data -----
        start_idx = None
        for idx, line in enumerate(lines):
            if line.strip().startswith("Hr"):
                start_idx = idx + 1
                break
        if start_idx is None:
            return cycle_str, location_str, None, None, effective_tz_name, "Data section not found in .bull file."

        m_old = re.search(r"(\d{8})\s*(\d{2})", cycle_str)
        cycle_date_str_old = date_str
        cycle_hour_str_old = run_str
        if m_old:
            cycle_date_str_old = m_old.group(1)
            cycle_hour_str_old = m_old.group(2)
        try:
            cycle_dt_utc_old = datetime.strptime(f"{cycle_date_str_old} {cycle_hour_str_old}", "%Y%m%d %H")
        except Exception:
            cycle_dt_utc_old = datetime.strptime(f"{date_str} {run_str}", "%Y%m%d %H")
        try:
            model_run_local_old = cycle_dt_utc_old.replace(tzinfo=UTC).astimezone(pytz.timezone(effective_tz_name))
        except Exception:
            model_run_local_old = cycle_dt_utc_old
        try:
            model_run_str = "Model Run: " + model_run_local_old.strftime("%A, %B %-d, %Y %I:%M %p")
        except Exception:
            model_run_str = "Model Run: " + model_run_local_old.strftime("%A, %B %d, %Y %I:%M %p").lstrip('0')

        for line in lines[start_idx:]:
            parts = line.split()
            if len(parts) < 20:
                continue
            try:
                hr_offset = float(parts[0])
            except ValueError:
                continue
            utc_dt = cycle_dt_utc_old + timedelta(hours=hr_offset)
            try:
                local_tz = pytz.timezone(effective_tz_name)
            except Exception:
                local_tz = UTC
            local_dt = utc_dt.replace(tzinfo=UTC).astimezone(local_tz)
            try:
                date_str_local = local_dt.strftime("%A, %B %-d, %Y")
            except Exception:
                date_str_local = local_dt.strftime("%A, %B %d, %Y").lstrip('0')
            time_str_local = local_dt.strftime("%I:%M %p").lstrip('0')
            row = [date_str_local, time_str_local]
            idx_base = 6
            for _swell in range(6):
                hs_val = tp_val = dir_val = None
                tokens_collected = 0
                while tokens_collected < 3 and idx_base < len(parts):
                    tok = parts[idx_base]
                    idx_base += 1
                    tok_clean = tok.replace('*', '')
                    if tok_clean == '':
                        continue
                    if tokens_collected == 0:
                        try:
                            hs_val = float(tok_clean) * 3.28084
                            tokens_collected += 1
                            continue
                        except ValueError:
                            continue
                    if tokens_collected == 1:
                        try:
                            tp_val = float(tok_clean)
                            tokens_collected += 1
                            continue
                        except ValueError:
                            continue
                    if tokens_collected == 2:
                        try:
                            dir_raw = int(float(tok_clean))
                            dir_val = (dir_raw + 180) % 360
                            tokens_collected += 1
                            continue
                        except ValueError:
                            continue
                if tokens_collected < 3:
                    row.extend([None, None, None])
                else:
                    row.extend([hs_val, tp_val, dir_val])
            combined_hs_ft = None
            for tok in reversed(parts):
                tok_clean = tok.replace('*', '')
                if tok_clean == '':
                    continue
                try:
                    combined_hs_ft = float(tok_clean) * 3.28084
                    break
                except ValueError:
                    continue
            row.append(combined_hs_ft)
            rows.append(row)

    # Round numeric values: Hs and Combined to 2 decimals, Tp to 1 decimal, Direction to int
    for i, r in enumerate(rows):
        idx_num = 2
        for _g in range(6):
            if r[idx_num] is not None:
                r[idx_num] = round(r[idx_num], 2)
            idx_num += 1
            if r[idx_num] is not None:
                r[idx_num] = round(r[idx_num], 1)
            idx_num += 1
            if r[idx_num] is not None:
                try:
                    r[idx_num] = int(round(r[idx_num]))
                except Exception:
                    pass
            idx_num += 1
        if r[-1] is not None:
            r[-1] = round(r[-1], 2)

    if not rows:
        return cycle_str, location_str, None, None, effective_tz_name, "No data rows parsed from .bull file."
    return cycle_str, location_str, model_run_str if 'model_run_str' in locals() else None, rows, effective_tz_name, None

# -----------------------------------------------------------------------------
# Table Formatting and Export Helpers
# -----------------------------------------------------------------------------
def build_html_table(cycle_str: str, location_str: str, model_run_str: str | None, rows: list[list], tz_label: str, unit: str) -> str:
    """
    Build an HTML table string mimicking the Excel "Table View" worksheet.
    """
    group_colors = [
        {"header": "#C00000", "subheader": "#F8B4B4", "data": "#F9DCDC"},
        {"header": "#ED7D31", "subheader": "#FBE5D6", "data": "#FDE7D4"},
        {"header": "#FFC000", "subheader": "#FFF2CC", "data": "#FFF9E5"},
        {"header": "#00B050", "subheader": "#D5E8D4", "data": "#EAF3E8"},
        {"header": "#00B0F0", "subheader": "#D9EAF6", "data": "#ECF5FB"},
        {"header": "#92D050", "subheader": "#E2F0D9", "data": "#F2F8EE"},
    ]
    combined_colors = {"header": "#7030A0", "subheader": "#D9D2E9", "data": "#EDE9F4"}
    total_cols = 2 + len(group_colors) * 3 + 1
    html = '<table class="table table-bordered table-sm">\n'
    html += f'<tr><td colspan="{total_cols}"><strong>{cycle_str}</strong></td></tr>\n'
    html += f'<tr><td colspan="{total_cols}"><strong>{location_str}</strong></td></tr>\n'
    # Model run is intentionally omitted in HTML
    html += f'<tr><td colspan="{total_cols}"><strong>Time Zone: {tz_label}</strong></td></tr>\n'
    html += '<tr>'
    html += '<th rowspan="2">Date</th>'
    html += '<th rowspan="2">Time</th>'
    for idx, col in enumerate(group_colors, start=1):
        html += f'<th colspan="3" style="background-color:{col["header"]}; color:white; text-align:center;">Swell {idx}</th>'
    html += f'<th style="background-color:{combined_colors["header"]}; color:white; text-align:center;">Combined</th>'
    html += '</tr>\n'
    html += '<tr>'
    hs_unit_label = '(ft)' if unit == 'US' else '(m)'
    for col in group_colors:
        html += f'<th style="background-color:{col["subheader"]}; text-align:center;">Hs<br>{hs_unit_label}</th>'
        html += f'<th style="background-color:{col["subheader"]}; text-align:center;">Tp<br>(s)</th>'
        html += f'<th style="background-color:{col["subheader"]}; text-align:center;">Dir<br>(d)</th>'
    html += f'<th style="background-color:{combined_colors["subheader"]}; text-align:center;">Hs<br>{hs_unit_label}</th>'
    html += '</tr>\n'
    for row in rows:
        try:
            parsed_time = datetime.strptime(row[1], "%I:%M %p").time()
        except Exception:
            try:
                parsed_time = datetime.strptime(row[1], "%I:%M:%S %p").time()
            except Exception:
                parsed_time = None
        bold_start = datetime.strptime("6:00:00 AM", "%I:%M:%S %p").time()
        bold_end = datetime.strptime("7:00:00 PM", "%I:%M:%S %p").time()
        dashed_start_evening = datetime.strptime("8:00:00 PM", "%I:%M:%S %p").time()
        dashed_end_morning = datetime.strptime("5:00:00 AM", "%I:%M:%S %p").time()
        border_style = ""
        font_weight_row = "normal"
        if parsed_time is not None:
            if bold_start <= parsed_time <= bold_end:
                border_style = "border:1px solid #000;"
                font_weight_row = "bold"
            elif parsed_time >= dashed_start_evening or parsed_time <= dashed_end_morning:
                border_style = "border:1px dashed #999;"
                font_weight_row = "normal"
        html += '<tr>'
        date_style = f'font-weight:bold; {border_style} padding:4px 8px;'
        html += f'<td style="{date_style}">{row[0]}</td>'
        time_style = f'font-weight:{font_weight_row}; {border_style} padding:4px 8px;'
        html += f'<td style="{time_style}">{row[1]}</td>'
        idx = 2
        for col in group_colors:
            val = row[idx]
            if val is None:
                hs_str = ""
            else:
                display_val = val if unit == 'US' else (val / 3.28084)
                hs_str = f"{display_val:.2f}"
            cell_style = f'background-color:{col["data"]}; text-align:right; font-weight:{font_weight_row}; {border_style} padding:4px 8px;'
            html += f'<td style="{cell_style}">{hs_str}</td>'
            idx += 1
            val = row[idx]
            tp_str = "" if val is None else f"{val:.1f}"
            html += f'<td style="{cell_style}">{tp_str}</td>'
            idx += 1
            val = row[idx]
            dir_str = "" if val is None else f"{val}"
            html += f'<td style="{cell_style}">{dir_str}</td>'
            idx += 1
        val = row[-1]
        if val is None:
            comb_str = ""
        else:
            display_comb = val if unit == 'US' else (val / 3.28084)
            comb_str = f"{display_comb:.2f}"
        cell_style = f'background-color:{combined_colors["data"]}; text-align:right; font-weight:{font_weight_row}; {border_style} padding:4px 8px;'
        html += f'<td style="{cell_style}">{comb_str}</td>'
        html += '</tr>\n'
    html += '</table>'
    return html

def build_excel_workbook(cycle_str: str, location_str: str, model_run_str: str | None, rows: list[list], tz_label: str, unit: str) -> BytesIO:
    """
    Create an Excel workbook that mirrors the formatting of the Excel "Table View".
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Table View"
    group_colors = [
        {"header": "C00000", "subheader": "F8B4B4", "data": "F9DCDC"},
        {"header": "ED7D31", "subheader": "FBE5D6", "data": "FDE7D4"},
        {"header": "FFC000", "subheader": "FFF2CC", "data": "FFF9E5"},
        {"header": "00B050", "subheader": "D5E8D4", "data": "EAF3E8"},
        {"header": "00B0F0", "subheader": "D9EAF6", "data": "ECF5FB"},
        {"header": "92D050", "subheader": "E2F0D9", "data": "F2F8EE"},
    ]
    combined_colors = {"header": "7030A0", "subheader": "D9D2E9", "data": "EDE9F4"}
    total_cols = 2 + len(group_colors) * 3 + 1
    row_idx = 1
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=total_cols)
    cell = ws.cell(row=row_idx, column=1, value=cycle_str)
    cell.font = Font(bold=True)
    row_idx += 1
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=total_cols)
    cell = ws.cell(row=row_idx, column=1, value=location_str)
    cell.font = Font(bold=True)
    row_idx += 1
    # Omit model run row in Excel per your current requirements
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=total_cols)
    cell = ws.cell(row=row_idx, column=1, value=f"Time Zone: {tz_label}")
    cell.font = Font(bold=True)
    row_idx += 1
    ws.cell(row=row_idx, column=1, value="Date").font = Font(bold=True)
    ws.cell(row=row_idx, column=2, value="Time").font = Font(bold=True)
    col_idx = 3
    for idx, colors in enumerate(group_colors, start=1):
        ws.merge_cells(start_row=row_idx, start_column=col_idx, end_row=row_idx, end_column=col_idx + 2)
        hdr_cell = ws.cell(row=row_idx, column=col_idx, value=f"Swell {idx}")
        hdr_cell.fill = PatternFill(start_color=colors["header"], end_color=colors["header"], fill_type="solid")
        hdr_cell.font = Font(bold=True, color="FFFFFF")
        hdr_cell.alignment = Alignment(horizontal="center")
        col_idx += 3
    comb_cell = ws.cell(row=row_idx, column=col_idx, value="Combined")
    comb_cell.fill = PatternFill(start_color=combined_colors["header"], end_color=combined_colors["header"], fill_type="solid")
    comb_cell.font = Font(bold=True, color="FFFFFF")
    comb_cell.alignment = Alignment(horizontal="center")
    row_idx += 1
    col_idx = 3
    hs_unit_label = "(ft)" if unit == 'US' else "(m)"
    for colors in group_colors:
        cell = ws.cell(row=row_idx, column=col_idx, value=f"Hs {hs_unit_label}")
        cell.fill = PatternFill(start_color=colors["subheader"], end_color=colors["subheader"], fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(bold=True)
        col_idx += 1
        cell = ws.cell(row=row_idx, column=col_idx, value="Tp (s)")
        cell.fill = PatternFill(start_color=colors["subheader"], end_color=colors["subheader"], fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(bold=True)
        col_idx += 1
        cell = ws.cell(row=row_idx, column=col_idx, value="Dir (d)")
        cell.fill = PatternFill(start_color=colors["subheader"], end_color=colors["subheader"], fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(bold=True)
        col_idx += 1
    cell = ws.cell(row=row_idx, column=col_idx, value=f"Hs {hs_unit_label}")
    cell.fill = PatternFill(start_color=combined_colors["subheader"], end_color=combined_colors["subheader"], fill_type="solid")
    cell.alignment = Alignment(horizontal="center")
    cell.font = Font(bold=True)
    for data_row in rows:
        row_idx += 1
        parsed_time = None
        try:
            parsed_time = datetime.strptime(data_row[1], "%I:%M:%S %p").time()
        except Exception:
            try:
                parsed_time = datetime.strptime(data_row[1], "%I:%M %p").time()
            except Exception:
                parsed_time = None
        bold_start = datetime.strptime("6:00:00 AM", "%I:%M:%S %p").time()
        bold_end = datetime.strptime("7:00:00 PM", "%I:%M:%S %p").time()
        dashed_start_evening = datetime.strptime("8:00:00 PM", "%I:%M:%S %p").time()
        dashed_end_morning = datetime.strptime("5:00:00 AM", "%I:%M:%S %p").time()
        is_bold_row = False
        border_style_name = None
        if parsed_time is not None:
            if bold_start <= parsed_time <= bold_end:
                is_bold_row = True
                border_style_name = "thin"
            elif parsed_time >= dashed_start_evening or parsed_time <= dashed_end_morning:
                is_bold_row = False
                border_style_name = "dashed"
        border_obj = None
        if border_style_name:
            border_obj = Border(
                left=Side(style=border_style_name, color="000000"),
                right=Side(style=border_style_name, color="000000"),
                top=Side(style=border_style_name, color="000000"),
                bottom=Side(style=border_style_name, color="000000"),
            )
        date_cell = ws.cell(row=row_idx, column=1, value=data_row[0])
        date_cell.font = Font(bold=True)
        if border_obj:
            date_cell.border = border_obj
        time_cell = ws.cell(row=row_idx, column=2, value=data_row[1])
        time_cell.font = Font(bold=is_bold_row)
        if border_obj:
            time_cell.border = border_obj
        col_idx = 3
        data_iter = iter(data_row[2:])
        for colors in group_colors:
            hs_val = next(data_iter)
            display_hs = hs_val
            if hs_val is not None and unit != 'US':
                display_hs = hs_val / 3.28084
            cell = ws.cell(row=row_idx, column=col_idx, value=display_hs if display_hs is not None else "")
            cell.fill = PatternFill(start_color=colors["data"], end_color=colors["data"], fill_type="solid")
            cell.number_format = "0.00"
            cell.font = Font(bold=is_bold_row)
            if border_obj:
                cell.border = border_obj
            col_idx += 1
            tp_val = next(data_iter)
            cell = ws.cell(row=row_idx, column=col_idx, value=tp_val if tp_val is not None else "")
            cell.fill = PatternFill(start_color=colors["data"], end_color=colors["data"], fill_type="solid")
            cell.number_format = "0.0"
            cell.font = Font(bold=is_bold_row)
            if border_obj:
                cell.border = border_obj
            col_idx += 1
            dir_val = next(data_iter)
            cell = ws.cell(row=row_idx, column=col_idx, value=dir_val if dir_val is not None else "")
            cell.fill = PatternFill(start_color=colors["data"], end_color=colors["data"], fill_type="solid")
            cell.font = Font(bold=is_bold_row)
            if border_obj:
                cell.border = border_obj
            col_idx += 1
        combined_val = data_row[-1]
        display_comb = combined_val
        if combined_val is not None and unit != 'US':
            display_comb = combined_val / 3.28084
        cell = ws.cell(row=row_idx, column=col_idx, value=display_comb if display_comb is not None else "")
        cell.fill = PatternFill(start_color=combined_colors["data"], end_color=combined_colors["data"], fill_type="solid")
        cell.number_format = "0.00"
        cell.font = Font(bold=is_bold_row)
        if border_obj:
            cell.border = border_obj
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    col = 3
    for _ in range(len(group_colors) * 3 + 1):
        letter = chr(64 + col) if col <= 26 else 'A' + chr(38 + col)  # simple extension for extra columns
        ws.column_dimensions[letter].width = 10
        col += 1
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


@app.route("/", methods=["GET", "POST"])
def index():
    """
    Home route: renders map + dropdowns, fetches selected station, shows table.
    """
    stations = get_station_list()
    timezones = sorted(pytz.common_timezones)
    unit_options = ["US", "Metric"]

    selected_station = ""
    selected_tz = ""
    selected_unit = "US"
    if request.method == "POST":
        selected_station = request.form.get("station") or ""
        selected_tz = request.form.get("tz") or ""
        selected_unit = request.form.get("unit") or "US"
    else:
        selected_station = request.args.get("station", "")
        selected_tz = request.args.get("tz", "")
        selected_unit = request.args.get("unit", "US") or "US"

    if not selected_station:
        selected_station = "51201"

    table_html = None
    error = None
    tz_label = ""
    selected_lat: float | None = None
    selected_lon: float | None = None

    if selected_station:
        cycle_str, location_str, model_run_str, rows, effective_tz_name, parse_error = parse_bull(selected_station, selected_tz or None)
        error = parse_error
        if rows is not None:
            tz_label = effective_tz_name
            table_html = build_html_table(cycle_str, location_str, model_run_str, rows, tz_label, selected_unit)
            coords_map = load_station_coords()
            sid_str = str(selected_station).strip()
            if sid_str in coords_map:
                selected_lat = coords_map[sid_str]['lat']
                selected_lon = coords_map[sid_str]['lon']
            else:
                if location_str:
                    m = re.search(r"\(\s*([-+]?\d+(?:\.\d+)?)\s*([NS])\s+([-+]?\d+(?:\.\d+)?)\s*([EW])\)", location_str)
                    if m:
                        try:
                            lat_val = float(m.group(1))
                            lat_dir = m.group(2).upper()
                            lon_val = float(m.group(3))
                            lon_dir = m.group(4).upper()
                            selected_lat = lat_val if lat_dir == 'N' else -lat_val
                            selected_lon = lon_val if lon_dir == 'E' else -lon_val
                        except Exception:
                            selected_lat = None
                            selected_lon = None

    return render_template(
        "index.html",
        stations=stations,
        selected_station=selected_station,
        timezones=timezones,
        selected_tz=selected_tz or tz_label,
        units=unit_options,
        selected_unit=selected_unit,
        table_html=table_html,
        error=error,
        selected_lat=selected_lat,
        selected_lon=selected_lon,
    )


@app.route("/download/<station_id>")
def download(station_id: str):
    """
    Download endpoint for Excel workbook.
    """
    tz_param = request.args.get("tz", "")
    unit_param = request.args.get("unit", "US") or "US"
    cycle_str, location_str, model_run_str, rows, effective_tz_name, error = parse_bull(station_id, tz_param or None)
    if rows is None:
        return f"Error: {error}", 404
    bio = build_excel_workbook(cycle_str, location_str, model_run_str, rows, effective_tz_name, unit_param)
    filename = f"{station_id}_table_view.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    app.run(debug=True)
